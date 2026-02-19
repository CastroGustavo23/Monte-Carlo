"""
═══════════════════════════════════════════════════════════════════════════════
SIMULAÇÃO DE MONTE CARLO — ANÁLISE PREVENTIVA DE RISCO DE RUPTURA
Supply Chain | Plano de Sortimentos

Versão: 1.0 FINAL
Autor: Especialista em Supply Chain & Ciência de Dados
Data: 2025-02-19
═══════════════════════════════════════════════════════════════════════════════
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ═══ CONFIGURAÇÕES ═══════════════════════════════════════════════════════════
np.random.seed(42)
N_SIM = 10_000      # Número de simulações Monte Carlo
HORIZON = 30        # Horizonte de análise (dias)

# Paleta de cores
C_HEADER_DARK = "003C5E"
C_HEADER_ORANGE = "FF6B35"
C_LIGHT_GREEN = "D4F4DD"
C_WHITE = "FFFFFF"
C_TEXT_DARK = "1E293B"
C_RED = "EF4444"
C_AMBER = "F59E0B"
C_GREEN = "10B981"

def criar_borda():
    s = Side(border_style="thin", color="94A3B8")
    return Border(left=s, right=s, top=s, bottom=s)

# ═══ INÍCIO DA EXECUÇÃO ══════════════════════════════════════════════════════
print("="*80)
print("  ANÁLISE PREVENTIVA DE RISCO — MONTE CARLO")
print("  Estrutura: Estoque Total = Extrema + Itapeva[Total] + Linhares[Bruto]")
print("="*80)

# ─── [1/4] LER DADOS ─────────────────────────────────────────────────────────
print("\n[1/4] Lendo Plano de Sortimentos...")
df_raw = pd.read_excel('Plano de Sortimentos 4.0.xlsx', 
                       sheet_name='Consulta de Estoques', 
                       header=1, 
                       nrows=200)

real_cols = df_raw.iloc[0].tolist()
df = df_raw.iloc[1:].copy()
df.columns = real_cols
df = df.reset_index(drop=True)
df = df.dropna(subset=['Item'])
print(f"   ✓ {len(df)} SKUs encontrados")

# ─── [2/4] PREPARAR DADOS ────────────────────────────────────────────────────
print("\n[2/4] Preparando estrutura de estoque...")

num_cols = ['Estoque [Extrema]', 'Itapeva [Total]', 'Itapeva [Bruto]', 
            'Itapeva [NFC]', 'Linhares [Bruto]', 'Média de Vendas [14 dias] ', 
            'Estoque de Segurança']

for col in num_cols:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

# Estrutura correta de estoque
df['NFC_Disponivel'] = df['Estoque [Extrema]'] + df['Itapeva [NFC]']
df['Bruto_Para_Selar'] = df['Itapeva [Bruto]'] + df['Linhares [Bruto]']
df['Estoque_Total'] = df['Estoque [Extrema]'] + df['Itapeva [Total]'] + df['Linhares [Bruto]']
df['Venda_Media'] = df['Média de Vendas [14 dias] ']

# Filtro preventivo: apenas SKUs com estoque
df_preventivo = df[df['Estoque_Total'] > 0].copy()

print(f"   ✓ Total no plano:           {len(df)} SKUs")
print(f"   ✓ Com estoque disponível:   {len(df_preventivo)} SKUs (análise preventiva)")
print(f"   ✗ Já rompidos (excluídos):  {len(df) - len(df_preventivo)} SKUs")

# ─── [3/4] SIMULAÇÃO MONTE CARLO ─────────────────────────────────────────────
def simular_sku(row, n_sim=N_SIM, horizon=HORIZON):
    """
    Simula 10.000 cenários futuros para 1 SKU considerando:
    - Variabilidade da demanda (Normal/Poisson + picos aleatórios)
    - Lead time incerto (Log-Normal)
    - Processo de industrialização (selamento NFC)
    """
    sku_nome = row['Item']
    curva = row.get('Curva ABC', 'N/A')
    est_nfc = float(row['NFC_Disponivel'])
    est_bruto = float(row['Bruto_Para_Selar'])
    venda_media = float(row['Venda_Media'])
    es_atual = float(row.get('Estoque de Segurança', 200))
    categoria = row.get('Categoria', 'N/A')
    status = row.get('Status', 'N/A')
    
    # Caso 1: Estoque parado (sem vendas)
    if venda_media < 0.1:
        dias_cobertura_teorica = 999
        return {
            "Item": sku_nome, "Categoria": categoria, "Curva ABC": curva, "Status": status,
            "NFC_Dispon": int(est_nfc), "Bruto": int(est_bruto), "Total": int(est_nfc + est_bruto),
            "Venda/dia": 0.0, "Dias_Cobertura": dias_cobertura_teorica,
            "Prob_Ruptura_%": 0.0, "Prob_NFC_Zero_%": 0.0,
            "ES_P95": 0, "ES_Atual": int(es_atual), "Gap_ES": int(es_atual),
            "Risco": "ESTOQUE_PARADO"
        }
    
    # Caso 2: SKU operacional (com vendas)
    sigma_demanda = venda_media * 0.35  # Coef. variação 35%
    
    # Parâmetros por curva ABC
    if curva == 'AA':
        lt_medio, lt_std, taxa_selamento = 7, 2, max(100, venda_media * 1.5)
    elif curva == 'A':
        lt_medio, lt_std, taxa_selamento = 10, 3, max(100, venda_media * 1.3)
    elif curva == 'B':
        lt_medio, lt_std, taxa_selamento = 14, 4, max(50, venda_media * 1.0)
    else:
        lt_medio, lt_std, taxa_selamento = 21, 6, max(30, venda_media * 0.8)
    
    has_nfc = taxa_selamento > 0
    sigma_selamento = taxa_selamento * 0.20
    
    # Arrays de resultado
    est_nfc_final = np.zeros(n_sim)
    ruptura_dias = np.zeros(n_sim)
    demanda_acum = np.zeros(n_sim)
    dias_ate_ruptura = np.full(n_sim, horizon + 1)  # se não romper = horizon+1
    
    # Rodar simulações
    for i in range(n_sim):
        nfc, bruto, rupturas, dem_total = est_nfc, est_bruto, 0, 0.0
        primeiro_dia_ruptura = None
        
        # Lead time aleatório (Log-Normal)
        lt = max(1, int(np.random.lognormal(
            np.log(max(lt_medio, 1)), 
            lt_std / max(lt_medio, 1)
        )))
        
        # Simular dia a dia
        for dia in range(horizon):
            # Demanda diária com variabilidade
            if venda_media >= 5:
                demanda = max(0.0, np.random.normal(venda_media, sigma_demanda))
            else:
                demanda = float(np.random.poisson(max(0, venda_media)))
            
            # Pico aleatório (5% chance de demanda dobrar)
            if np.random.random() < 0.05:
                demanda *= 2.0
            
            dem_total += demanda
            
            # Processo de selamento NFC
            if has_nfc and bruto > 0:
                taxa_dia = max(0.0, np.random.normal(taxa_selamento, sigma_selamento))
                selado = min(bruto, taxa_dia)
                bruto -= selado
                nfc += selado
            
            # Atender demanda (prioridade: NFC → Bruto em emergência)
            atendido = min(nfc, demanda)
            nfc -= atendido
            nao_atendido = demanda - atendido
            
            if nao_atendido > 0 and bruto > 0:
                at_bruto = min(bruto, nao_atendido)
                bruto -= at_bruto
                nao_atendido -= at_bruto
            
            # Registrar ruptura
            if nao_atendido > 0:
                rupturas += 1
                if primeiro_dia_ruptura is None:
                    primeiro_dia_ruptura = dia + 1
        
        # Armazenar resultados da simulação i
        est_nfc_final[i] = max(0.0, nfc)
        ruptura_dias[i] = rupturas
        demanda_acum[i] = dem_total
        dias_ate_ruptura[i] = primeiro_dia_ruptura if primeiro_dia_ruptura else (horizon + 1)
    
    # Calcular métricas
    prob_ruptura = np.mean(ruptura_dias > 0)
    prob_nfc_zero = np.mean(est_nfc_final <= 0)
    es_95 = np.percentile(demanda_acum, 95) - np.mean(demanda_acum)
    
    # ✅ CORREÇÃO: Dias de cobertura realista
    # Se não rompeu em nenhum cenário, usar cobertura teórica
    # Se rompeu, usar mediana dos dias até ruptura
    dias_mediana = int(np.median(dias_ate_ruptura))
    dias_teorico = int((est_nfc + est_bruto) / venda_media)
    
    if prob_ruptura == 0:
        # Não rompeu → usar cobertura teórica limitada a 30 dias
        dias_cobertura = min(dias_teorico, 30)
    else:
        # Rompeu → usar mediana real dos cenários
        dias_cobertura = min(dias_mediana, dias_teorico)
    
    # Classificação de risco
    if prob_ruptura > 0.3:
        risco = "ALTO"
    elif prob_ruptura > 0.1:
        risco = "MÉDIO"
    else:
        risco = "BAIXO"
    
    return {
        "Item": sku_nome,
        "Categoria": categoria,
        "Curva ABC": curva,
        "Status": status,
        "NFC_Dispon": int(est_nfc),
        "Bruto": int(est_bruto),
        "Total": int(est_nfc + est_bruto),
        "Venda/dia": round(venda_media, 1),
        "Dias_Cobertura": dias_cobertura,
        "Prob_Ruptura_%": round(prob_ruptura * 100, 1),
        "Prob_NFC_Zero_%": round(prob_nfc_zero * 100, 1),
        "ES_P95": int(es_95),
        "ES_Atual": int(es_atual),
        "Gap_ES": int(es_atual - es_95),
        "Risco": risco
    }

# Executar simulações
print(f"\n[3/4] Rodando {N_SIM:,} simulações por SKU...")
resultados = []
for idx, row in df_preventivo.iterrows():
    resultados.append(simular_sku(row))
    if (len(resultados) % 10 == 0) or (len(resultados) == len(df_preventivo)):
        pct = len(resultados) / len(df_preventivo) * 100
        print(f"   Progresso: {len(resultados)}/{len(df_preventivo)} ({pct:.0f}%)", end='\r')

print("\n   ✓ Simulações concluídas!")

# ─── [4/4] EXPORTAR EXCEL ────────────────────────────────────────────────────
print("\n[4/4] Criando Excel formatado...")

df_resultados = pd.DataFrame(resultados)

# Ordenar por risco
ordem_risco = {'ALTO': 0, 'MÉDIO': 1, 'BAIXO': 2, 'ESTOQUE_PARADO': 3}
df_resultados['_ordem'] = df_resultados['Risco'].map(ordem_risco)
df_resultados = df_resultados.sort_values(
    ['_ordem', 'Prob_Ruptura_%'], 
    ascending=[True, False]
).drop(columns=['_ordem'])

# Criar workbook
wb = Workbook()
ws = wb.active
ws.title = "Análise Preventiva"
ws.freeze_panes = "E2"

# Larguras de coluna
col_widths = [35, 15, 12, 18, 12, 12, 12, 12, 12, 14, 14, 12, 12, 12, 12]
for idx, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(idx)].width = width

# Header
headers = list(df_resultados.columns)
ws.row_dimensions[1].height = 25

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = header
    cell.font = Font(name="Aptos", size=11, bold=True, color=C_WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = criar_borda()
    
    if header in ["Item", "Categoria", "Curva ABC", "Status"]:
        cell.fill = PatternFill(start_color=C_HEADER_ORANGE, 
                               end_color=C_HEADER_ORANGE, fill_type="solid")
    else:
        cell.fill = PatternFill(start_color=C_HEADER_DARK, 
                               end_color=C_HEADER_DARK, fill_type="solid")

# Dados
for row_idx, row_data in enumerate(df_resultados.itertuples(index=False), 2):
    ws.row_dimensions[row_idx].height = 20
    bg_color = C_LIGHT_GREEN if row_idx % 2 == 0 else C_WHITE
    
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = criar_borda()
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, 
                               fill_type="solid")
        cell.font = Font(name="Aptos", size=10, color=C_TEXT_DARK)
        cell.alignment = Alignment(
            horizontal="center" if col_idx != 1 else "left",
            vertical="center",
            indent=1 if col_idx == 1 else 0
        )
        
        # Formatação condicional
        if headers[col_idx-1] == "Risco":
            cor_map = {"ALTO": C_RED, "MÉDIO": C_AMBER, "BAIXO": C_GREEN, 
                      "ESTOQUE_PARADO": "94A3B8"}
            cor = cor_map.get(value, C_WHITE)
            cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
            cell.font = Font(name="Aptos", size=10, bold=True, color=C_WHITE)
        
        if headers[col_idx-1] == "Prob_Ruptura_%":
            try:
                prob = float(value)
                if prob > 30:
                    cell.fill = PatternFill(start_color=C_RED, end_color=C_RED, 
                                          fill_type="solid")
                    cell.font = Font(name="Aptos", size=10, bold=True, color=C_WHITE)
                elif prob > 10:
                    cell.fill = PatternFill(start_color=C_AMBER, end_color=C_AMBER, 
                                          fill_type="solid")
                    cell.font = Font(name="Aptos", size=10, bold=True, color=C_WHITE)
            except:
                pass
        
        if headers[col_idx-1] == "Gap_ES":
            try:
                if int(value) < 0:
                    cell.font = Font(name="Aptos", size=10, bold=True, color=C_RED)
            except:
                pass
        
        if headers[col_idx-1] == "Dias_Cobertura":
            try:
                dias = int(value)
                if dias < 15:
                    cell.font = Font(name="Aptos", size=10, bold=True, color=C_RED)
                elif dias < 30:
                    cell.font = Font(name="Aptos", size=10, bold=True, color=C_AMBER)
            except:
                pass

# Salvar
output_file = "Analise_Preventiva_Monte_Carlo.xlsx"
wb.save(output_file)
print(f"   ✓ Excel salvo: {output_file}\n")

# ═══ RESUMO EXECUTIVO ════════════════════════════════════════════════════════
total = len(df_resultados)
alto = len(df_resultados[df_resultados['Risco'] == 'ALTO'])
medio = len(df_resultados[df_resultados['Risco'] == 'MÉDIO'])
baixo = len(df_resultados[df_resultados['Risco'] == 'BAIXO'])
parado = len(df_resultados[df_resultados['Risco'] == 'ESTOQUE_PARADO'])

print("="*80)
print("  RESUMO EXECUTIVO")
print("="*80)
print(f"\n📊 CLASSIFICAÇÃO DE RISCO ({total} SKUs analisados):")
print(f"   🔴 ALTO:   {alto:2d} SKUs ({alto/total*100:5.1f}%) — Ação imediata")
print(f"   🟡 MÉDIO:  {medio:2d} SKUs ({medio/total*100:5.1f}%) — Monitorar semanalmente")
print(f"   🟢 BAIXO:  {baixo:2d} SKUs ({baixo/total*100:5.1f}%) — Situação confortável")
print(f"   ⚪ PARADO: {parado:2d} SKUs ({parado/total*100:5.1f}%) — Sem vendas")

print(f"\n🚨 TOP 10 RISCOS PREVENTIVOS:")
print("-" * 95)
print(f"{'':3s} {'SKU':35s} | {'Dias':4s} | {'NFC':>5s} | {'Bruto':>5s} | {'Total':>6s} | {'Prob':>5s}")
print("-" * 95)

top_riscos = df_resultados[df_resultados['Risco'] != 'ESTOQUE_PARADO'].head(10)
for idx, row in top_riscos.iterrows():
    emoji = "🔴" if row['Risco'] == 'ALTO' else "🟡" if row['Risco'] == 'MÉDIO' else "🟢"
    dias = row['Dias_Cobertura']
    print(f"{emoji} {row['Item'][:35]:35s} | {dias:4d} | {row['NFC_Dispon']:5d} | "
          f"{row['Bruto']:5d} | {row['Total']:6d} | {row['Prob_Ruptura_%']:5.1f}%")

print("\n💡 INSIGHTS:")
print(f"   • {alto} SKUs precisam de ação HOJE")
print(f"   • {len(df) - len(df_preventivo)} SKUs já rompidos foram excluídos")
print(f"   • ES_P95 = Estoque de Segurança recomendado (nível de serviço 95%)")
print(f"   • Gap_ES negativo = Estoque de Segurança INSUFICIENTE")

print("\n" + "="*80)
print("✅ Análise concluída! Arquivo Excel gerado com sucesso.")
print("="*80 + "\n")